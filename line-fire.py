from __future__ import unicode_literals
from datetime import datetime, timezone, timedelta

import random
import sqlite3
import datetime
import errno
import json
import os
import sys
import tempfile
import requests
from argparse import ArgumentParser
from dotenv import load_dotenv

from flask import Flask, request, abort, send_from_directory,make_response,send_file,url_for,flash,render_template,redirect
from flask_paginate import Pagination, get_page_args


from werkzeug.middleware.proxy_fix import ProxyFix

from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.exceptions import (
    LineBotApiError, InvalidSignatureError
)
from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage,
    SourceUser, PostbackEvent, StickerMessage, StickerSendMessage,FlexSendMessage, 
    LocationMessage, LocationSendMessage, ImageMessage, ImageSendMessage,FileMessage)

import time
from pathlib import Path

import cv2
import torch
import torch.backends.cudnn as cudnn
from numpy import random

from models.experimental import attempt_load
from utils.datasets import LoadStreams, LoadImages
from utils.general import check_img_size, non_max_suppression, apply_classifier, scale_coords, xyxy2xywh, \
    strip_optimizer, set_logging, increment_path
from utils.plots import plot_one_box
from utils.torch_utils import select_device, load_classifier, time_synchronized
import pickle 
import wikipedia
import pyrebase
import datetime
import openpyxl
import wolframalpha

from dateutil.parser import parse

#======== function check if string is datetime 


def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    try: 
        parse(string, fuzzy=fuzzy)
        return True

    except ValueError:
        return False

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_host=1, x_proto=1)
app.secret_key = "Secret Key"

# reads the key-value pair from .env file and adds them to environment variable.
load_dotenv()

# get channel_secret and channel_access_token from your environment variable
channel_secret = os.getenv('LINE_CHANNEL_SECRET', None)
channel_access_token = os.getenv('LINE_CHANNEL_ACCESS_TOKEN', None)
if channel_secret is None or channel_access_token is None:
    print('Specify LINE_CHANNEL_SECRET and LINE_CHANNEL_ACCESS_TOKEN as environment variables.')
    sys.exit(1)

line_bot_api = LineBotApi(channel_access_token)
handler = WebhookHandler(channel_secret)

static_tmp_path = os.path.join(os.path.dirname(__file__), 'static', 'tmp')
    
#===== Firebase ====================================================
config = {
    'apiKey': "AIzaSyCl-wWBByv3261Uu_h5IBqp15fV5iMaGzI",
    'authDomain': "ai-detection-5a2b6.firebaseapp.com",
    'databaseURL': "https://ai-detection-5a2b6-default-rtdb.firebaseio.com",
    'projectId': "ai-detection-5a2b6",
    'storageBucket': "ai-detection-5a2b6.appspot.com",
    'messagingSenderId': "865732945786",
    'appId': "1:865732945786:web:e671ea10f3ecff9c82d8c3",
    'measurementId': "G-4FJ30SDMPG",
  };


firebase = pyrebase.initialize_app(config)
auth = firebase.auth()

user = auth.sign_in_with_email_and_password("kevinangas@gmail.com", "Mikey131998")


# profile = line_bot_api.get_profile(user_id)

storage = firebase.storage()




def add_image_db(img_path):
    tz = timezone(timedelta(hours = 7))
    now = datetime.datetime.now(tz=tz)
    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    path_on_cloud = f"images/{timestamp}.jpg"
    path_local = img_path
    storage.child(path_on_cloud).put(path_local)


#========================================= Admin Panel ==========================================


@app.route("/register",methods=['GET'])
def register(): 
    
    return render_template("register.html")


@app.route("/register_success",methods=['GET','POST'])
def register_success(): 
    user_id = request.form['username'].strip()
    sex = request.form['sex']
    age = request.form['age']
    weight =  request.form['weight']
    height = request.form['height']
    exercise_rate = request.form['exercise_rate']
    
    with sqlite3.connect("fooddb.db") as con: 
        curr = con.cursor()

        sql_cmd = """
        
        INSERT INTO USER VALUES(?,?,?,?,?,?);
        
        """

        curr.execute(sql_cmd,(user_id,sex,age,weight,height,exercise_rate))
        con.commit()
    
    flash("Register Successfully")
    return redirect(url_for("register"))
    

@app.route('/admin')
def Index():

    
    with sqlite3.connect("fooddb.db") as con: 

        curr = con.cursor()
        sql_cmd = """
        SELECT user_id,food_detect,date,portion FROM UserFood; 
        
        """
        users =  [ (i[0],i[1],i[2],i[3],line_bot_api.get_profile(i[0]).display_name) for i in curr.execute(sql_cmd)]


    def get_users(offset=0, per_page=10):
        

        return users[offset: offset + per_page]

    """
    all_data = [ (i[0],i[1],i[2],i[3],line_bot_api.get_profile(i[0]).display_name) for i in curr.execute(sql_cmd)]

    """
    
    page, per_page, offset = get_page_args(page_parameter='page',
                                           per_page_parameter='per_page')
    total = len(users)


    
    print(per_page)
    pagination_users = get_users(offset, per_page)
    print(pagination_users)
    pagination = Pagination(page=page, per_page=per_page , total=total,
                            css_framework='bootstrap5')


    
    
    return render_template("index.html",employees=pagination_users,
                           page=page,
                           per_page=per_page,
                           pagination=pagination)
 


 
#this route is for inserting data to mysql database via html forms

 
 
#this is our update route where we are going to update our employee
@app.route('/update', methods = ['GET', 'POST'])
def update():
    
    if request.method == 'POST':

        with sqlite3.connect("fooddb.db") as con: 

            curr = con.cursor()

 
            user_id = request.form['user_id']
            date = request.form['date']
            # food_detect = request.form['food_detect']
            portion = request.form['portion']


            # Update ============

        portion_input = portion
        date_input = date

        with sqlite3.connect('fooddb.db') as con:
            
            curr = con.cursor()
            sql_cmd = """
            
            SELECT * FROM Food;
            """
            curr.execute(sql_cmd)
            food_nut_detail = []
            for i in curr.fetchall(): 
                food_nut_detail.append(i)
        
        print(food_nut_detail)

        with sqlite3.connect("fooddb.db") as con: 
            curr = con.cursor()

            sql_cmd = """
            
            SELECT * FROM USERFOOD WHERE user_id == ? and date == ?;
            
            """

            food = curr.execute(sql_cmd,(user_id,date_input))

            food_name_portion = []
            sum_protein = 0 
            sum_carbo = 0 
            sum_fat = 0 
            
            
        """
        Update row sumfat sumcarbo sumprotein
     
     
        """
        protein = []
        carbo = []
        fat = []
        energy = []
        with sqlite3.connect("fooddb.db") as con: 
            curr = con.cursor()
        
            sql_cmd = """
                    
                    SELECT * FROM USERFOOD WHERE user_id == ? and date == ?;
                    
                    """
        
         
            
            
            food = curr.execute(sql_cmd,(user_id,date_input)).fetchall() 
            food_name_portion = []
            print(portion_input)
            food_portion = [float(r) for r in portion_input.split()]
            food_detect_name = food[0][1].split(",")
            

            
            for food_name in food_detect_name: 
                for row in food_nut_detail: 
                    if food_name == row[0]: 
                        protein.append(row[1])
                        fat.append(row[2])
                        carbo.append(row[3])
                        energy.append(row[4])
            
            portion_protein = zip(protein,food_portion)
            portion_carbo = zip(carbo,food_portion)
            portion_fat = zip(fat,food_portion)
            portion_energy = zip(energy,food_portion) #Energy

            
        sum_protein = sum([float(cal)*float(portion) for cal,portion in portion_protein])
        sum_carbo = sum([float(cal)*float(portion) for cal,portion in portion_carbo])

        sum_fat = sum([float(cal)*float(portion) for cal,portion in portion_fat])
        sum_energy = sum([float(cal)*float(portion) for cal,portion in portion_energy]) #Energy
            
            # for i in food_nut_detail:
            #     for j in range(len(food_detect_name)): 
            #         if i[0] == food_detect_name[j]:  #i[1]=protein i[2]=fat i[3]=carbo
            #             print(float(i[1]))
            #             print(float(i[2]))
            #             print(float(i[3]))
            #             protein.append( float(i[1]) * float(food_portion[j] ) )
            #             fat.append( float(i[2]) * float(food_portion[j] ) )
            #             carbo.append( float(i[3]) * float(food_portion[j] ) )
            

        with sqlite3.connect("fooddb.db") as con: 
                                
             curr = con.cursor()

             sql_cmd = """
                                
             UPDATE USERFOOD SET portion = ?,protein=?,fat=?,carbo=?,energy=? WHERE user_id == ? and date == ?;  

             """

             portion_insert = portion_input
             print(portion_insert)

             curr.execute(sql_cmd ,(portion_insert ,sum_protein ,sum_fat,sum_carbo,sum_energy ,str(user_id),date_input)) 



            #==================
            
            

            
        flash("Updated Successfully") 
 
        return redirect(url_for('Index'))
 
 
 
 
#This route is for deleting our employee
@app.route('/delete', methods = ['GET', 'POST'])
def delete():
    with sqlite3.connect("fooddb.db") as con: 
        curr = con.cursor()
        sql_cmd = """
                                
             DELETE FROM USERFOOD WHERE user_id == ? and date == ?; 

             """
 
        user_id = request.form['user_id']
        date = request.form['date']
        curr.execute(sql_cmd,(user_id,date)) 


    flash("Users Deleted Successfully")
 
    return redirect(url_for('Index'))
 
 
 





#=================================================Admin Edit User Information==============================================

@app.route('/edit_user')
def Index2():

    
    with sqlite3.connect("fooddb.db") as con: 

        curr = con.cursor()
        sql_cmd = """
        SELECT * FROM User;
        
        """
        users =  [ (i[0],i[1],i[2],i[3],i[4],i[5]) for i in curr.execute(sql_cmd)]


    def get_users(offset=0, per_page=10):
        

        return users[offset: offset + per_page]

    """
    all_data = [ (i[0],i[1],i[2],i[3],line_bot_api.get_profile(i[0]).display_name) for i in curr.execute(sql_cmd)]

    """
    
    page, per_page, offset = get_page_args(page_parameter='page',
                                           per_page_parameter='per_page')
    total = len(users)


    
    print(per_page)
    pagination_users = get_users(offset, per_page)
    print(pagination_users)
    pagination = Pagination(page=page, per_page=per_page , total=total,
                            css_framework='bootstrap5')


    
    
    return render_template("userinfo.html",employees=pagination_users,
                           page=page,
                           per_page=per_page,
                           pagination=pagination)
 


 
#this route is for inserting data to mysql database via html forms

 
 
#this is our update route where we are going to update our employee
@app.route('/update2', methods = ['GET', 'POST'])
def update2():
    
    if request.method == 'POST':

        with sqlite3.connect("fooddb.db") as con: 

            curr = con.cursor()

            userid = request.form['id']
            sex = request.form['sex']
            age = request.form['age']
            weight = request.form['weight']
            height = request.form['height']
            ex_rate = request.form['ex_rate']
 
            sql_cmd = """
            
            UPDATE USER SET sex=?,age=?,weight=?,height=?,exercise_rate=? WHERE userid==?;
            """

            curr.execute(sql_cmd,(sex,age,weight,height,ex_rate,userid))

            # Update ============





            
        flash("Updated Successfully") 
 
    return redirect(url_for('Index2'))
 
 
 
 
#=======================================



### YOLOv5 ###
# Setup
def delete_all(): 
    try: 
        for i in os.listdir("static"): 
            os.remove(i)
    except: 
        pass

# function for create tmp dir for download content
def make_static_tmp_dir():
    try:
        os.makedirs(static_tmp_path)
    except OSError as exc:
        if exc.errno == errno.EEXIST and os.path.isdir(static_tmp_path):
            pass
        else:
            raise

@app.route("/profile/<userid>",methods=['GET'])
def get_profile(userid): 
    profile = line_bot_api.get_profile(userid)
    profile_name = profile.display_name
    
    return profile_name  


@app.route("/callback/download",methods=['GET'])
def download_food_result():
    try:
        global filename_excel 
    
        return send_file(filename_excel,as_attachment=True,cache_timeout=0) # ใส่เพื่อไม่ให้มันจำไฟ์เดิม cache_timeout=0
    except:
        print("user_id_download is null")
        

@app.route("/", methods=['GET'])
def home():

    return "<marquee behavior='scroll' direction='left' scrollamount=15>Food Recognition AI BMR and Test</marquee>"



@app.route("/callback", methods=['POST'])
def callback():
    # get X-Line-Signature header value
    signature = request.headers['X-Line-Signature']

    # get request body as text
    body = request.get_data(as_text=True)
    app.logger.info("Request body: " + body)

    # handle webhook body
    try:
        handler.handle(body, signature)
    except LineBotApiError as e:
 
        print("Got exception from LINE Messaging API: %s\n" % e.message)
        for m in e.error.details:
            print("  %s: %s" % (m.property, m.message))
        print("\n")
   
     
    except InvalidSignatureError:
        abort(400)

    return 'OK'



@handler.add(MessageEvent, message=TextMessage)
def handle_text_message(event):
    #เรียกใช้ฟังก์ชัน generate_answer เพื่อแยกส่วนของคำถาม
    text = event.message.text
    if text.lower() == "1": 
            weights = "yolov5s.pt"
            with open("model.pkl","wb") as f: 
                f.write(pickle.dumps(weights))
            line_bot_api.reply_message(
                event.reply_token, [
                    TextSendMessage(text="เปลี่ยนเป็นโหมดตรวจจับวัตถุ"),
                ]
            )       
    elif text.lower() == "2": 
            weights = "food.pt"
            with open("model.pkl","wb") as f: 
                f.write(pickle.dumps(weights))
            line_bot_api.reply_message(
                event.reply_token, [
                    TextSendMessage(text="เปลี่ยนเป็นโหมดตรวจจับอาหารไทย"),
                ]
            )           
    elif text.lower() == "3":
            weights = "best_BCCM.pt"
            with open("model.pkl","wb") as f: 
                f.write(pickle.dumps(weights))
            line_bot_api.reply_message(
                event.reply_token, [
                    TextSendMessage(text="เปลี่ยนเป็นโหมดตรวจจับเม็ดเลือด"),
                ]
            )     

    elif "คำนวณ" in text.lower():
        try:

            client = wolframalpha.Client("G84Y4Q-6K6GVKXLRA")
           
        
            response = client.query(text[text.index("คำนวณ")+len("คำนวณ "):])
            
            result = None
            for result in response.results:
                pass
            # You could have also simply used result = list(response.results)[-1]
            
            if result is not None:
                ans_real = f"คำตอบของคุณ คือ {result.text}".format(result.text)
                
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text=ans_real))
        except:
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ไม่เข้าใจเลยค่ะ พูดใหม่ได้ไหม"))  
    elif text.lower() == "user": 
        
        userID = event.source.user_id # เป็นการรับค่า  user_id จากคนอื่น
        text_result = "{} {}".format(event.source.user_id,event.message.text)
        line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text=userID))    
    
    elif text.lower() == "เปลี่ยนเป็นภาษาอังกฤษ" or text.lower() == "english" or text.lower() == "eng" or text.lower() == "อังกฤษ" or text.lower() == "ค้นหาด้วยาษาอังกฤษ":
        wikipedia.set_lang('en')
        line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ค้นหา wikipedia ด้วยภาษาอังกฤษ"))
    elif text.lower() == "เปลี่ยนเป็นภาษาไทย" or text.lower() == "thai" or text.lower() == "th" or text.lower() == "ไทย" or text.lower() == "ค้นหาด้วยาษาไทย":
        wikipedia.set_lang('th')
        line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ค้นหา wikipedia ด้วยภาษาไทย"))
    elif text.lower() == "delete" or text.lower() == "ลบภาพ" or text.lower() == "ลบ":
        for i in os.listdir(static_tmp_path): 
            os.remove(os.path.join(static_tmp_path,i))
        line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ลบภาพทั้งหมด เรียบร้อยแล้วค่ะ"))
    elif "ค้นหา" in text.lower():
        try:
            ans = wikipedia.summary(text[text.index("ค้นหา")+len("ค้นหา"):])
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text=ans))
        except:
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ไม่เข้าใจเลยค่ะ พูดใหม่ได้ไหม"))
    elif "search" in text.lower(): 
        text = text.lower()
   
        try:
            
            ans = wikipedia.summary(text[text.index("search")+len("search"):])
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text=ans))

        except:
            line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ไม่เข้าใจเลยค่ะ พูดใหม่ได้ไหม"))
    
    elif "ตั้งชื่อบอท" in text.lower():
        text = text.lower() 
        global botname  
        botname = text[text.index("ตั้งชื่อบอท")+len("ตั้งชื่อบอท"):]
        line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="ขอบคุณสำหรับการตั้งชื่อให้ {} นะคะ".format(botname)))

    elif is_date(" ".join(text.lower().split(" ")[:2])): 
        
        portion_input = text.lower().split(" ")[2:]
        date_input = " ".join(text.lower().split(" ")[:2])

        with sqlite3.connect('fooddb.db') as con:
            
            curr = con.cursor()
            sql_cmd = """
            
            SELECT * FROM Food;
            """
            curr.execute(sql_cmd)
            food_nut_detail = []
            for i in curr.fetchall(): 
                food_nut_detail.append(i)
        
        print(food_nut_detail)

        with sqlite3.connect("fooddb.db") as con: 
            curr = con.cursor()

            sql_cmd = """
            
            SELECT * FROM USERFOOD WHERE user_id == ? and date == ?;
            
            """

            food = curr.execute(sql_cmd,(event.source.user_id,date_input))

            food_name_portion = []
            sum_protein = 0 
            sum_carbo = 0 
            sum_fat = 0 
            
            
        """
        Update row sumfat sumcarbo sumprotein
     
     
        """
        protein = []
        carbo = []
        fat = []
        energy = []
        with sqlite3.connect("fooddb.db") as con: 
            curr = con.cursor()
        
            sql_cmd = """
                    
                    SELECT * FROM USERFOOD WHERE user_id == ? and date == ?;
                    
                    """
        
         
            user_id = event.source.user_id
            
            food = curr.execute(sql_cmd,(user_id,date_input)).fetchall() 
            food_name_portion = []
   
            
            food_portion = [float(r) for r in portion_input]
            food_detect_name = food[0][1].split(",")
            

            
            for food_name in food_detect_name: 
                for row in food_nut_detail: 
                    if food_name == row[0]: 
                        protein.append(row[1])
                        fat.append(row[2])
                        carbo.append(row[3])
                        energy.append(row[4])
            
            portion_protein = zip(protein,food_portion)
            portion_carbo = zip(carbo,food_portion)
            portion_fat = zip(fat,food_portion)
            portion_energy = zip(energy,food_portion) #Energy

            
        sum_protein = sum([float(cal)*float(portion) for cal,portion in portion_protein])
        sum_carbo = sum([float(cal)*float(portion) for cal,portion in portion_carbo])

        sum_fat = sum([float(cal)*float(portion) for cal,portion in portion_fat])
        sum_energy = sum([float(cal)*float(portion) for cal,portion in portion_energy]) #Energy
            
            # for i in food_nut_detail:
            #     for j in range(len(food_detect_name)): 
            #         if i[0] == food_detect_name[j]:  #i[1]=protein i[2]=fat i[3]=carbo
            #             print(float(i[1]))
            #             print(float(i[2]))
            #             print(float(i[3]))
            #             protein.append( float(i[1]) * float(food_portion[j] ) )
            #             fat.append( float(i[2]) * float(food_portion[j] ) )
            #             carbo.append( float(i[3]) * float(food_portion[j] ) )
            

        with sqlite3.connect("fooddb.db") as con: 
                                
             curr = con.cursor()

             sql_cmd = """
                                
             UPDATE USERFOOD SET portion = ?,protein=?,fat=?,carbo=?,energy=? WHERE user_id == ? and date == ?;  

             """

             portion_insert = " ".join(portion_input)
             print(portion_insert)

             curr.execute(sql_cmd ,(portion_insert ,sum_protein ,sum_fat,sum_carbo,sum_energy ,str(event.source.user_id),date_input)) 
            



        line_bot_api.reply_message(
                event.reply_token, TextSendMessage(text="อัปเดตสัดส่วนอาหาร"))
    

    
    elif "รายการอาหาร" in text.lower():  # รายการอาหาร วันที่ เพศ อายุ น้ำหนัก ส่วนสูง

            with sqlite3.connect("fooddb.db") as con: 
                curr = con.cursor()
                sql_cmd = """
                
                SELECT * FROM USER; 
                """
                user_lis = [i for i in curr.execute(sql_cmd)]
            print([i[0] for i in user_lis])
            print([i[0] for i in user_lis].index(event.source.user_id))
            if event.source.user_id in [i[0] for i in user_lis]: 
                for i in user_lis: 
                     print(i[0])
                     if event.source.user_id == i[0]: 
                        user_input_date = text.split(" ")[1]
                        sex = i[1]
                        age = i[2]
                        weight = i[3]
                        height = i[4]
                        ex_rate = i[5]

            else:
                user_input_date = text.split(" ")[1]
                sex = text.split(" ")[2]
                age = text.split(" ")[3]
                weight = text.split(" ")[4]
                height = text.split(" ")[5]
                ex_rate = int(text.split(" ")[6])
            #user_input_date = text[text.index("รายการอาหาร")+len("รายการอาหาร"):]
            
            # line_bot_api.reply_message(
            #     event.reply_token, TextSendMessage(text=user_input_date))
            
            
            if sex == "ชาย": 
                bmr = 65.5+9.56* float(weight) + 1.85*float(height) - 4.68*float(age)
                bmr_result = "BMR ชาย : {} kcal".format(bmr)
            elif sex == "หญิง": 
                bmr = 66.5+13.75*float(weight) + 5*float(height) - 6.78*float(age)
                bmr_result = "BMR หญิง : {} kcal".format(bmr)
            # daily_cal
            # daily_cal_result
            if ex_rate == 0: 
                tdee = round(bmr*1.2,2)
                tdee_result = "TDEE เท่ากับ {} kcal".format(tdee)
            elif ex_rate>=1 and ex_rate <=2: 
                tdee = round(bmr*1.375,2)
                tdee_result = "TDEE เท่ากับ {} kcal".format(tdee)
            elif ex_rate>=3 and ex_rate <=5: 
                tdee = round(bmr*1.55,2)
                tdee_result = "TDEE เท่ากับ {} kcal".format(tdee)
            elif ex_rate>=6 and ex_rate <=7: 
                tdee = round(bmr*1.725,2)
                tdee_result = "TDEE เท่ากับ {} kcal".format(tdee)
            elif ex_rate >7: 
                tdee = round(bmr*1.9,2)
                tdee_result = "TDEE เท่ากับ {} kcal".format(tdee)                
                    
            try:
                wb = openpyxl.Workbook()
                sheet = wb.worksheets[0]
             
                with sqlite3.connect("fooddb.db") as con: 
                                
                                curr = con.cursor()
                                user_input = event.source.user_id
                                date_input = user_input_date
                                sql_cmd = """
                                
                                SELECT * FROM USERFOOD WHERE USER_ID == ?; 
                                """
                                index_count = 1
                                curr.execute(sql_cmd,(user_input,))
                                
                                sheet.cell(index_count,1).value = "ลำดับที่" # ใส่ลงช่องที่กำหนด
                                sheet.cell(index_count,2).value = "รายชื่ออาหาร" # ใส่ลงช่องที่กำหนด    
                                sheet.cell(index_count,3).value = "โปรตีนรวม (กรัม)" # ใส่ลงช่องที่กำหนด 
                                sheet.cell(index_count,4).value = "ไขมันรวม (กรัม)" # ใส่ลงช่องที่กำหนด 
                                sheet.cell(index_count,5).value = "คาร์โบไฮเดรตรวม (กรัม)" # ใส่ลงช่องที่กำหนด 
                                sheet.cell(index_count,6).value = "พลังงานรวม (kcal/จาน)" # ใส่ลงช่องที่กำหนด 
                                sheet.cell(index_count,7).value = "วันที่" # ใส่ลงช่องที่กำหนด 
                                sheet.cell(index_count,8).value = "สัดส่วนอาหาร (จาน) ตามลำดับ" # ใส่ลงช่องที่กำหนด 
                                
                                sum_energy_day = 0
                                sum_protein_day = 0 
                                sum_fat_day = 0 
                                sum_carbo_day = 0
                        
                                for i in curr.fetchall(): 
                                    # print(i[6].split(" ")[0])
                                    
                                    if date_input.strip() == i[6].split(" ")[0]:
                                        # print(i)
                                        #wb.append((index_count,i[1],i[2],i[3],i[4],i[5]))
                                        
                                        sum_protein_day = sum_protein_day + i[2]
                                        sum_fat_day = sum_fat_day + i[3]
                                        sum_carbo_day = sum_carbo_day + i[4]
                                        sum_energy_day = sum_energy_day + i[5]
                                        
                                        sheet.cell(index_count+1,1).value = index_count # ใส่ลงช่องที่กำหนด
                                        sheet.cell(index_count+1,2).value = i[1] # ใส่ลงช่องที่กำหนด    
                                        sheet.cell(index_count+1,3).value = i[2] # ใส่ลงช่องที่กำหนด 
                                        sheet.cell(index_count+1,4).value = i[3] # ใส่ลงช่องที่กำหนด 
                                        sheet.cell(index_count+1,5).value = i[4] # ใส่ลงช่องที่กำหนด 
                                        sheet.cell(index_count+1,6).value = i[5] # ใส่ลงช่องที่กำหนด 
                                        sheet.cell(index_count+1,7).value = i[6] # ใส่ลงช่องที่กำหนด 
                                        sheet.cell(index_count+1,8).value = i[7] # ใส่ลงช่องที่กำหนด 
                                        index_count = index_count + 1
                                        
                                    
                                   
                                            #======== generate excel from databse ============
                
                                
                                sheet.cell(index_count+1,1).value = "โปรตีนรวมต่อวัน (กรัม)"
                                sheet.cell(index_count+1,2).value = sum_protein_day
                                
                                sheet.cell(index_count+2,1).value = "ไขมันรวมต่อวัน (กรัม)"
                                sheet.cell(index_count+2,2).value = sum_fat_day
                                
                                sheet.cell(index_count+3,1).value = "คาร์โบไฮเดรตรวมต่อวัน (กรัม)"
                                sheet.cell(index_count+3,2).value = sum_carbo_day
                                
                                sheet.cell(index_count+4,1).value = "พลังงานรวมต่อวัน (kcal)"
                                sheet.cell(index_count+4,2).value = sum_energy_day
                                
                                
                                """
                                
                                แนะนำโปรตีน คาร์โบ ไขมัน พลังงาน เกินไม่เกิน
                                """
                                if sum_protein_day >= 45 and sum_protein_day <= 60: 
                                    protein_result = "โปรตีนระดับปกติ {} kcal".format(round(sum_protein_day,2))
                                elif sum_protein_day < 45: 
                                    protein_result = "โปรตีนต่ำกว่าปกติ {} kcal".format(round(sum_protein_day,2))
                                elif sum_protein_day >60: 
                                    protein_result = "โปรตีนมากกว่าปกติ {} kcal".format(round(sum_protein_day,2))
                                    
                                    
                                if sum_fat_day >= 500 and sum_fat_day <=700: 
                                    fat_result = "ไขมันระดับปกติ {} kcal".format(round(sum_fat_day,2))
                                elif sum_fat_day < 500: 
                                    fat_result = "ไขมันระดับต่ำกว่าปกติ {} kcal".format(round(sum_fat_day,2))
                                elif sum_fat_day >700: 
                                    fat_result = "ไขมันระดับมากกว่าปกติ {} kcal".format(round(sum_fat_day,2))
                                
                                
                                if sum_carbo_day <= 65: 
                                    carbo_result = "คาร์โบไฮเดรตระดับปกติ {} kcal".format(round(sum_carbo_day,2))
                                else: 
                                    carbo_result = "คาร์โบไฮเดรตมากกว่าปกติ {} kcal".format(round(sum_carbo_day,2))
                                
                                
                                global filename_excel 
                                user_id_download = event.source.user_id
                                filename_excel = str(user_id_download)  + ".xlsx"
                                wb.save(filename_excel)
                                # x = "\n".join([str(i) for i in result_food_list]) 
                                # print(len(x))
    
           
                if sum_energy_day > tdee :
                    cal_desc_day = "ได้รับปริมาณแคลอรีต่อวันเพียงพอแล้ว ({} kcal)".format(round(sum_energy_day,2))
                elif sum_energy_day < tdee: 
                    cal_desc_day = "ขาดปริมาณแคลอรีต่อวันอีก ({} kcal)".format(round(tdee-sum_energy_day,2))
                    
                imgUrl = "https://supapongai.com/wp-content/uploads/2021/09/Medium-minifile.gif"
                desc = "{}\n{}\n{}\n{}\n{}\n{}".format(tdee_result,bmr_result,protein_result,fat_result,carbo_result,cal_desc_day)
                
                TitleName = "สรุปรายการอาหารประจำวัน"
              
                
                flex = """
                 {
            "type": "bubble",
            "hero": {
              "type": "image",
              "url": "%s",
              "margin": "none",
              "size": "full",
              "aspectRatio": "1:1",
              "aspectMode": "cover",
              "action": {
                "type": "uri",
                "label": "Action",
                "uri": "https://linecorp.com"
              }
            },
            "body": {
              "type": "box",
              "layout": "vertical",
              "spacing": "md",
              "action": {
                "type": "uri",
                "label": "Action",
                "uri": "https://linecorp.com"
              },
              "contents": [
                {
                  "type": "text",
                  "text": "%s",
                  "size": "xl",
                  "weight": "bold"
                },
                {
                  "type": "text",
                  "text": "%s",
                  "wrap": true
                }
              ]
            },
            "footer": {
              "type": "box",
              "layout": "vertical",
              "contents": [
                {
                  "type": "button",
                  "action": {
                    "type": "uri",
                    "label": "รายการอาหารประจำวัน",
                    "uri": "%s"
                  },
                  "color": "#F67878",
                  "style": "primary"
                }
              ]
            }
            }
                
                """ %(imgUrl,TitleName,desc, request.url+"/download")
                
                """
                Flexbox send message ต้องแปลงเป็น json ก่อน แล้วต้องใส่ strict=False ไม่งั้นเกิด JsonDecodeError  
                """
                flex = json.loads(flex,strict=False)
                replyObj = FlexSendMessage(alt_text='Flex Message alt text', contents=flex) 
                
    
                
                
                line_bot_api.reply_message(event.reply_token, [replyObj])
                
                """
                line_bot_api.reply_message(
                                 event.reply_token, [TextSendMessage(text="{}\n{}\n{}\n{}\n{}".format(protein_result,fat_result,carbo_result,request.url+"/download",date_input.strip()))])                        
                """
            except:
                line_bot_api.reply_message(
                    event.reply_token, TextSendMessage(text="ไม่เข้าใจเลยค่ะ พูดใหม่ได้ไหม"))

    
    else: 
       
        try:
            
            user_id2 = event.source.user_id
            profile = line_bot_api.get_profile(user_id2)
            user_name = profile.display_name
            url = "https://openapi.botnoi.ai/service-api/botnoichitchat?keyword={}&styleid=11&botname={}".format(text,botname)
            headers = {
                    'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2MjgxOTAyNzUsImlkIjoiYTMxYzkyM2QtYzkxMy00YTU3LWE5OTEtYjMzZGFjNjZhZWJlIiwiaXNzIjoiSXEzNHdwdmpWMWNydUs1NDNkUVZQZHoxM2JkU1BsdTUiLCJuYW1lIjoiU3VwYXBvbmciLCJwaWMiOiJodHRwczovL3Byb2ZpbGUubGluZS1zY2RuLm5ldC8waFRVSm82RzZCQzNwM0ZpUXBlREIwTFV0VEJSY0FPQTB5RDNkRVQxVWVCUmdJZGtrcFFuRVFHZ1VYWFVNUGRrbDVHU0pHSGxRUlZVMEsifQ.AB2WDl5RTqrWC2TNRDrmVHlyG1AivSDseCc7cWJ0PHc'
                }
            response = requests.request("GET", url, headers=headers)
            result_text = json.loads(response.text)['reply']
            line_bot_api.reply_message(
                        event.reply_token, TextSendMessage(text=result_text)) 
        except NameError: 
           
            botname = "สวีตตี้"
            user_id2 = event.source.user_id
            profile = line_bot_api.get_profile(user_id2)
            user_name = profile.display_name
            url = "https://openapi.botnoi.ai/service-api/botnoichitchat?keyword={}&styleid=11&botname={}".format(text,botname)
            headers = {
                    'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE2MjgxOTAyNzUsImlkIjoiYTMxYzkyM2QtYzkxMy00YTU3LWE5OTEtYjMzZGFjNjZhZWJlIiwiaXNzIjoiSXEzNHdwdmpWMWNydUs1NDNkUVZQZHoxM2JkU1BsdTUiLCJuYW1lIjoiU3VwYXBvbmciLCJwaWMiOiJodHRwczovL3Byb2ZpbGUubGluZS1zY2RuLm5ldC8waFRVSm82RzZCQzNwM0ZpUXBlREIwTFV0VEJSY0FPQTB5RDNkRVQxVWVCUmdJZGtrcFFuRVFHZ1VYWFVNUGRrbDVHU0pHSGxRUlZVMEsifQ.AB2WDl5RTqrWC2TNRDrmVHlyG1AivSDseCc7cWJ0PHc'
                }
            response = requests.request("GET", url, headers=headers)
            result_text = json.loads(response.text)['reply']
            line_bot_api.reply_message(
                        event.reply_token, TextSendMessage(text=result_text))          



@handler.add(MessageEvent, message=LocationMessage)
def handle_location_message(event):
    line_bot_api.reply_message(
        event.reply_token,
        LocationSendMessage(
            title='Location', address=event.message.address,
            latitude=event.message.latitude, longitude=event.message.longitude
        )
    )


@handler.add(MessageEvent, message=StickerMessage)
def handle_sticker_message(event):
    line_bot_api.reply_message(
        event.reply_token,
        StickerSendMessage(
            package_id=event.message.package_id,
            sticker_id=event.message.sticker_id)
    )




# Other Message Type
@handler.add(MessageEvent, message=(ImageMessage))
def handle_content_message(event):
    #============== Set Up Yolo============================
    
    with open('model.pkl','rb') as f: 
        
        a = pickle.loads(f.read())
        
    weights, view_img, save_txt, imgsz = a, False, False, 640
    
    # weights, view_img, save_txt, imgsz = 'yolov5s.pt', False, False, 640
    conf_thres = 0.25
    iou_thres = 0.45
    classes = None
    agnostic_nms = False
    save_conf = False
    save_img = True
    
    # Directories
    save_dir = 'static/tmp/'
    
    # Initialize
    set_logging()
    device = select_device('')
    half = device.type != 'cpu'  # half precision only supported on CUDA
    
    # Load model
    model = attempt_load(weights, map_location=device)  # load FP32 model
    imgsz = check_img_size(imgsz, s=model.stride.max())  # check img_size
    if half:
        model.half()  # to FP16
        
    if isinstance(event.message, ImageMessage):
        ext = 'jpg'
    else:
        return
    

    
    
    message_content = line_bot_api.get_message_content(event.message.id)
    with tempfile.NamedTemporaryFile(dir=static_tmp_path, prefix=ext + '-', delete=False) as tf:
        for chunk in message_content.iter_content():
            tf.write(chunk)
        tempfile_path = tf.name

    dist_path = tempfile_path + '.' + ext
    
    

    dist_name = os.path.basename(dist_path)
    os.rename(tempfile_path, dist_path)
    add_image_db(dist_path)
    # Set Dataloader
    dataset = LoadImages(dist_path, img_size=imgsz)
        
    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names
    colors = [[random.randint(0, 255) for _ in range(3)] for _ in names]

    # Run inference
    t0 = time.time()
    img = torch.zeros((1, 3, imgsz, imgsz), device=device)  # init img
    _ = model(img.half() if half else img) if device.type != 'cpu' else None  # run once
    object_label = []
    for path, img, im0s, vid_cap in dataset:
        img = torch.from_numpy(img).to(device)
        img = img.half() if half else img.float()  # uint8 to fp16/32
        img /= 255.0  # 0 - 255 to 0.0 - 1.0
        if img.ndimension() == 3:
            img = img.unsqueeze(0)

        # Inference
        t1 = time_synchronized()
    
        pred = model(img, augment=False)[0]

        # Apply NMS
        pred = non_max_suppression(pred, conf_thres, iou_thres, classes=classes, agnostic=agnostic_nms)
        t2 = time_synchronized()

        # Process detections
        for i, det in enumerate(pred):  # detections per image
            p, s, im0, frame = path, '', im0s, getattr(dataset, 'frame', 0)

            p = Path(p)  # to Path
            save_path = str(save_dir + p.name)  # img.jpg
            s += '%gx%g ' % img.shape[2:]  # print string
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            if len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_coords(img.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, -1].unique():
                    n = (det[:, -1] == c).sum()  # detections per class
                    s += f'{n} {names[int(c)]}s, '  # add to string

                # Write results
                for *xyxy, conf, cls in reversed(det):
                    if save_txt:  # Write to file
                        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format

                    if save_img or view_img:  # Add bbox to image
                        label = f'{names[int(cls)]} {conf:.2f}'
                        
                        object_label.append(plot_one_box(xyxy, im0, label=label, color=colors[int(cls)], line_thickness=3))

            # Save results (image with detections)
            if save_img:
                if dataset.mode == 'image':
                    cv2.imwrite(save_path, im0)

            # Print time (inference + NMS)
            print(f'{s}Done. ({t2 - t1:.3f}s)')
            
    url = request.url_root + 'static/tmp/' + dist_name


    with sqlite3.connect('fooddb.db') as con:
        
        curr = con.cursor()
        sql_cmd = """
        
        SELECT * FROM Food;
        """
        curr.execute(sql_cmd)
        food_nut_detail = []
        for i in curr.fetchall(): 
            food_nut_detail.append(i)
    
    print(food_nut_detail)
    food_detect = object_label
    object_label2= [i[0:-4].strip() for i in food_detect]
    print(object_label2)
    
    with sqlite3.connect("fooddb.db") as con: 
            
        user_id = event.source.user_id
        
        nut_detail = []
        tz = timezone(timedelta(hours = 7))
        year = datetime.datetime.now(tz=tz).year
        month = datetime.datetime.now(tz=tz).month
        day = datetime.datetime.now(tz=tz).day 
        hour = datetime.datetime.now(tz=tz).hour
        minute =  datetime.datetime.now(tz=tz).minute
        date_send_now = "{}/{}/{} {}:{}".format(day,month,year,hour,minute)
                
        print(date_send_now.split(" ")[0])  
        for food_detect_name in object_label2:
            
            for food_nut in food_nut_detail: 
                    # food_nut[0] = food_name 
                if food_detect_name == food_nut[0].strip():
                        
                    protein = food_nut[1]
                    fat = food_nut[2]
                    carbo = food_nut[3]
                    energy = food_nut[4]
                    nut_detail.append( ( food_detect_name,protein,fat,carbo,energy ) ) 
            
            
            #============= Database Insert Section =======================
        curr = con.cursor()
        sql_cmd = """
            
            INSERT INTO USERFOOD VALUES(?,?,?,?,?,?,?,?);
            """
            
        sum_energy_day = sum([i[4] for i in nut_detail])
        sum_protein = sum([i[1] for i in nut_detail])  
        sum_fat = sum([i[2] for i in nut_detail])  
        sum_carbo = sum([i[3] for i in nut_detail])  
        real_portion = [str(1) for i in range(len(object_label))]

        print(sum_energy_day)
        print(sum_protein)
        print(sum_fat)
        print(sum_carbo)
         
        print(nut_detail)
        
        curr.execute(sql_cmd,(user_id,",".join(object_label2), sum_protein , sum_fat, sum_carbo,sum_energy_day , date_send_now," ".join(real_portion))) # nut_detail error 
        con.commit()




    line_bot_api.reply_message(
        event.reply_token, [
            TextSendMessage(text="รายงานผลลัพธ์การตรวจจับวัตถุ : \n พบวัตถุจำนวน {}  ชิ้น \n {}".format(len(object_label),",".join(object_label2))),
            ImageSendMessage(url,url)
        ])

@app.route('/static/<path:path>')
def send_static_content(path):
    return send_from_directory('static', path)

# create tmp dir for download content
make_static_tmp_dir()

if __name__ == "__main__":
    app.run(port=8000, debug=True)

