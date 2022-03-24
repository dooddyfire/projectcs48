# -*- coding: utf-8 -*-
"""
Created on Fri Oct  1 23:17:24 2021

@author: PREDATOR
"""
import sqlite3
import datetime 
import openpyxl 
#========= 

# with sqlite3.connect('fooddb.db') as con:
    
#     curr = con.cursor()
#     sql_cmd = """
    
#     SELECT * FROM Food;
#     """
#     curr.execute(sql_cmd)
#     food_nut_detail = []
#     for i in curr.fetchall(): 
#         food_nut_detail.append(i)

# print(food_nut_detail)
# food_detect =  ['Basil Rice 0.56', 'Basil Rice 0.59', 'Basil Rice 0.88']
# object_label2= [i[0:-4].strip() for i in food_detect]
# print(object_label2)

# with sqlite3.connect("fooddb.db") as con: 
        
#     user_id = "player001"
    
#     nut_detail = []
#     year = datetime.datetime.now().year
#     month = datetime.datetime.now().month
#     day = datetime.datetime.now().day 
#     hour = datetime.datetime.now().hour
#     minute =  datetime.datetime.now().minute
#     date_send_now = "{}/{}/{} {}:{}".format(day,month,year,hour,minute)
            
#     print(date_send_now.split(" ")[0])  
#     for food_detect_name in object_label2:
        
#         for food_nut in food_nut_detail: 
#                 # food_nut[0] = food_name 
#             if food_detect_name == food_nut[0].strip():
                    
#                 protein = food_nut[1]
#                 fat = food_nut[2]
#                 carbo = food_nut[3]
#                 energy = food_nut[4]
#                 nut_detail.append( ( food_detect_name,protein,fat,carbo,energy ) ) 
        
        
#         #============= Database Insert Section =======================
#     curr = con.cursor()
#     sql_cmd = """
        
#         INSERT INTO USERFOOD VALUES(?,?,?,?,?,?,?);
#         """
        
#     sum_energy_day = sum([i[4] for i in nut_detail])
#     sum_protein = sum([i[1] for i in nut_detail])  
#     sum_fat = sum([i[2] for i in nut_detail])  
#     sum_carbo = sum([i[3] for i in nut_detail])  
#     print(sum_energy_day)
#     print(sum_protein)
#     print(sum_fat)
#     print(sum_carbo)
     
#     print(nut_detail)
    
#     curr.execute(sql_cmd,(user_id,",".join(object_label2), sum_protein , sum_fat, sum_carbo,sum_energy_day , date_send_now)) # nut_detail error 
#     con.commit()

# with sqlite3.connect("fooddb.db") as con: 
    
#     curr = con.cursor()
#     user_input = "Uc277ddba165a4335a4ba39d4718c7693"
#     date_input = "2/10/2021"
#     sql_cmd = """
    
#     SELECT * FROM USERFOOD WHERE USER_ID == ?; 
#     """
    
#     curr.execute(sql_cmd,(user_input,))
    
#     for i in curr.fetchall(): 
#         # print(i[6].split(" ")[0])
#         if date_input == i[6].split(" ")[0]:
#             print(i)

wb = openpyxl.Workbook()
sheet = wb.worksheets[0]

with sqlite3.connect("fooddb.db") as con: 
                
                curr = con.cursor()
                user_input = "Uc277ddba165a4335a4ba39d4718c7693"
                date_input = "2/10/2021"
                sql_cmd = """
                
                SELECT * FROM USERFOOD WHERE USER_ID == ?; 
                """
                index_count = 1
                curr.execute(sql_cmd,(user_input,))
                
                sheet.cell(index_count,1).value = "ลำดับที่" # ใส่ลงช่องที่กำหนด
                sheet.cell(index_count,2).value = "รายชื่ออาหาร" # ใส่ลงช่องที่กำหนด    
                sheet.cell(index_count,3).value = "โปรตีนรวม" # ใส่ลงช่องที่กำหนด 
                sheet.cell(index_count,4).value = "ไขมันรวม" # ใส่ลงช่องที่กำหนด 
                sheet.cell(index_count,5).value = "คาร์โบไฮเดรตรวม" # ใส่ลงช่องที่กำหนด 
                sheet.cell(index_count,6).value = "พลังงานรวม (Kcal/จาน)" # ใส่ลงช่องที่กำหนด 
                sheet.cell(index_count,7).value = "วันที่" # ใส่ลงช่องที่กำหนด 
                
                
                
                
                for i in curr.fetchall(): 
                    # print(i[6].split(" ")[0])
                    
                    if date_input == i[6].split(" ")[0]:
                        # print(i)
                        #wb.append((index_count,i[1],i[2],i[3],i[4],i[5]))

                        sheet.cell(index_count+1,1).value = index_count # ใส่ลงช่องที่กำหนด
                        sheet.cell(index_count+1,2).value = i[1] # ใส่ลงช่องที่กำหนด    
                        sheet.cell(index_count+1,3).value = i[2] # ใส่ลงช่องที่กำหนด 
                        sheet.cell(index_count+1,4).value = i[3] # ใส่ลงช่องที่กำหนด 
                        sheet.cell(index_count+1,5).value = i[4] # ใส่ลงช่องที่กำหนด 
                        sheet.cell(index_count+1,6).value = i[5] # ใส่ลงช่องที่กำหนด 
                        sheet.cell(index_count+1,7).value = i[6] # ใส่ลงช่องที่กำหนด 
                        index_count = index_count + 1
                        
                    
                   
                            #======== generate excel from databse ============

                
#                            raw_data.append(i)
#                            num = num + 1
     
                wb.save("result_food.xlsx")
                # x = "\n".join([str(i) for i in result_food_list]) 
                # print(len(x))
